"""Tests automatisés d'exactitude des données — Chatbot ELISFA.

Lit la feuille 'Exactitude données' du plan xlsx et interroge le bot
sur chaque question de référence, puis vérifie que la réponse contient
bien les marqueurs attendus (valeurs numériques, articles de loi, etc.).

Usage :
    # Test en local (assume Flask lancé sur http://localhost:5000)
    python3 tests/run_data_accuracy.py

    # Test en production (URL publique)
    python3 tests/run_data_accuracy.py --url https://felias-reseau-eli2026.duckdns.org

    # Filtrer une catégorie
    python3 tests/run_data_accuracy.py --filter SSC

    # Écrire rapport JSON
    python3 tests/run_data_accuracy.py --report report.json

    # Mettre à jour la colonne 'Statut' + 'Dernière vérif.' directement dans le xlsx
    python3 tests/run_data_accuracy.py --update-xlsx

À planifier : cron mensuel (1er du mois, 3h du matin).
    0 3 1 * * cd /app && python3 tests/run_data_accuracy.py --update-xlsx
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
import unicodedata
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erreur : openpyxl requis. pip install openpyxl", file=sys.stderr)
    sys.exit(1)

try:
    import urllib.request
    import urllib.error
except ImportError:
    print("Erreur : module urllib indisponible.", file=sys.stderr)
    sys.exit(1)


# ── Configuration par défaut ──────────────────────────────────────────────
DEFAULT_URL = "http://localhost:5000"
DEFAULT_XLSX = Path(__file__).parent / "plan_evaluation_elisfa.xlsx"
DEFAULT_SHEET = "2. Exactitude données"
TIMEOUT_S = 60
PAUSE_BETWEEN_CALLS_S = 0.8  # Respect du rate-limit
MODULE_PAR_CATEGORIE = {
    "Rémunération : SSC": "juridique",
    "Rémunération : Valeur du point": "juridique",
    "Classification : ancienneté": "juridique",
    "Classification : critères classants": "juridique",
    "Classification : emplois repères": "juridique",
    "Droit du travail : articles Code": "juridique",
    "Droit du travail : articles CCN": "juridique",
    "Formation professionnelle": "formation",
    "Gouvernance associative": "gouvernance",
    "RH stratégique": "rh",
    "Anti-hallucination": "juridique",
    "Cohérence de nommage": "juridique",
}


# ── Utilitaires ───────────────────────────────────────────────────────────
def _strip_accents(s: str) -> str:
    """Supprime les accents pour comparaison tolérante."""
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    )


def _normalize(s: str) -> str:
    """Normalise une chaîne : minuscules, sans accents, espaces multiples réduits."""
    s = _strip_accents(s.lower())
    s = re.sub(r"\s+", " ", s).strip()
    return s


def _contains(haystack: str, needle: str) -> bool:
    """Vérifie la présence d'un marqueur, robuste aux variations typographiques.

    - Normalise les espaces insécables, fines, etc.
    - Ignore les accents et la casse.
    - Accepte des variantes de séparateurs numériques (22 100, 22100, 22.100).
    """
    # Normalise les espaces (insécable, fine, etc.) vers espace simple
    h = re.sub(r"[\u00A0\u202F\u2009\s]+", " ", haystack)
    n = re.sub(r"[\u00A0\u202F\u2009\s]+", " ", needle)
    h_norm = _normalize(h)
    n_norm = _normalize(n)
    if n_norm in h_norm:
        return True
    # Variante numérique : accepte 22100 / 22.100 / 22,100 si on cherche "22 100"
    if re.match(r"^\d[\d\s.,]*\d$", n.strip()):
        digits_only = re.sub(r"[^\d]", "", n)
        h_digits = re.sub(r"[^\d]", " ", h)
        if digits_only and digits_only in re.sub(r"\s+", "", h_digits):
            return True
    return False


# ── Parsing des critères de validation ────────────────────────────────────
# Les règles sont exprimées en français lisible dans la colonne "Critère de
# validation". On reconnaît plusieurs patterns français courants :
#   - "contient 'X' et 'Y'"            → must_contain_all
#   - "Réponse = X"                    → must_contain_all
#   - "Points exacts : 0, 5, 15, ..."  → must_contain_all
#   - "Mini X, maxi Y"                 → must_contain_all
#   - "cite L1234-5"                   → must_contain_all
#   - "mentionne X € / N %"            → must_contain_all
#   - "NE PAS inventer / dire 'X'"     → must_not_contain
# Les cas restants remontent comme "MANUEL".

# Quotes français : exige que la chaîne ouvre après un non-apostrophe (début/espace/deux-points)
# et ne dépasse pas ~60 chars, pour éviter de confondre les apostrophes typographiques
# avec des délimiteurs (« l'article », « d'application »).
RE_QUOTED = re.compile(
    r"(?:^|[\s:=])'([^'\n;]{1,60})'|"
    r'"([^"\n;]{1,60})"|'
    r"«\s*([^»\n;]{1,60})\s*»"
)
# Retire les parenthèses explicatives type "(espaces/points acceptés)"
RE_PARENS = re.compile(r"\(([^)]*)\)")
# Articles de loi : L1234-5, R1234-5, D1234-5, article 12, etc.
RE_ARTICLE = re.compile(r"\b([LRDA]\d{3,5}(?:-\d{1,3})?)\b")
# Montants : 22 100 € / 2 355 € / 1 261
RE_MONNAIE = re.compile(r"\b(\d[\d\s\u00A0\u202F.,]*)\s*(?:€|euros?|k€)\b", re.IGNORECASE)
# Points en liste : "Points exacts : 0/SSC, 5, 15, 30, 45, 65, 80, 110"
# Accepte des items comme "0/SSC" en première position.
RE_POINTS_LIST = re.compile(
    r"(?:points?\s+(?:exacts?|pr[ée]cis)|valeurs?)\s*[:=]\s*([\d/A-Z,\s]+(?:,\s*\d+)+)",
    re.IGNORECASE,
)
# "Mini X, maxi Y" — tolère des unités entre (pt, pts, points, €, heures…)
RE_MINI_MAXI = re.compile(
    r"mini\s*(?:mum)?\s*[=:]?\s*(\d+[\d\s.,]*)\s*(?:pts?|points?|€|euros?|heures?|h)?[,.;\s]+\s*maxi\s*(?:mum)?\s*[=:]?\s*(\d+[\d\s.,]*)",
    re.IGNORECASE,
)
# Pourcentages : "3 000 €", "50%", "50 %"
RE_POURCENT = re.compile(r"\b(\d+[\d\s.,]*)\s*%")
# Seuils heures : "17h30", "6 mois", "2 ans"
RE_DUREE = re.compile(r"\b(\d+)\s*(h\d{0,2}|mois|ans?|jours?|semaines?)", re.IGNORECASE)
# Dates : "2024", "01/01/2024", "2024-2027"
RE_ANNEE = re.compile(r"\b(20\d{2})\b")
RE_DATE_FR = re.compile(r"\b(\d{1,2}/\d{1,2}/\d{4})\b")

NEG_MARKERS = [
    re.compile(r"ne\s+(?:doit|devrait)\s+pas\s+(?:inventer|mentionner|contenir|dire)", re.I),
    re.compile(r"\bNE\s+PAS\s+(?:inventer|répondre|dire)", re.I),
    re.compile(r"aucune?\s+(?:hallucination|invention)", re.I),
    re.compile(r"sans\s+(?:inventer|halluciner)", re.I),
]
POS_MARKERS = [
    re.compile(r"contient|contenir|inclut|mentionne|cite|liste", re.I),
    re.compile(r"r[ée]ponse\s*[=:]", re.I),
    re.compile(r"\br[ée]ponse\s+(?:attendue|exacte)", re.I),
]


def _clean_parens(s: str) -> str:
    """Retire les parenthèses explicatives (non-normatives)."""
    return re.sub(RE_PARENS, "", s).strip()


def parse_criteria(texte: str) -> dict:
    """Extrait des contrôles structurés depuis un critère en français.

    Retourne un dict :
        {
            "must_contain_all": [m1, m2, ...],
            "must_contain_any": [[m1, m2], ...],  # au moins un par sous-liste
            "must_not_contain": [m1, m2, ...],
            "raw": "<texte original>",
            "parseable": bool,
        }
    """
    out = {
        "must_contain_any": [],
        "must_contain_all": [],
        "must_not_contain": [],
        "raw": texte,
        "parseable": False,
    }
    if not texte:
        return out

    texte_clean = _clean_parens(texte)

    # 1) Chaînes entre quotes : marqueurs de haute confiance
    for m in RE_QUOTED.finditer(texte):
        val = (m.group(1) or m.group(2) or m.group(3) or "").strip().rstrip(",.;:!?")
        # Rejette les faux positifs de type "l'article" (apostrophe typographique)
        if val and len(val) < 150 and val not in out["must_contain_all"]:
            out["must_contain_all"].append(val)

    # 2) Articles du Code du travail (L2251-1, L1234-1, etc.)
    for m in RE_ARTICLE.finditer(texte_clean):
        art = m.group(1)
        if art not in out["must_contain_all"]:
            out["must_contain_all"].append(art)

    # 3) Liste "Points exacts : 0, 5, 15, ..."
    m_pts = RE_POINTS_LIST.search(texte_clean)
    if m_pts:
        raw_list = m_pts.group(1)
        values = [v.strip() for v in re.split(r"[,/]", raw_list) if v.strip()]
        # Filtre : uniquement les nombres purs
        values = [v for v in values if re.match(r"^\d+$", v)]
        if len(values) >= 3:
            for v in values:
                if v not in out["must_contain_all"]:
                    out["must_contain_all"].append(v)

    # 4) Mini/Maxi
    m_mm = RE_MINI_MAXI.search(texte_clean)
    if m_mm:
        mini = m_mm.group(1).strip().rstrip(",.;:")
        maxi = m_mm.group(2).strip().rstrip(",.;:")
        if mini and mini not in out["must_contain_all"]:
            out["must_contain_all"].append(mini)
        if maxi and maxi not in out["must_contain_all"]:
            out["must_contain_all"].append(maxi)

    # 5) Montants en euros (22 100 €, 55 €, 3 000 €, 28 265 €)
    for m in RE_MONNAIE.finditer(texte_clean):
        val = m.group(1).strip().rstrip(".,")
        if val and val not in out["must_contain_all"]:
            out["must_contain_all"].append(val)

    # 6) Durées (17h30, 6 mois, 2 ans, 4 ans)
    for m in RE_DUREE.finditer(texte_clean):
        dur = f"{m.group(1)} {m.group(2)}".strip()
        # Normaliser "17h30" sans espace
        if re.match(r"^\d+h\d+$", m.group(2)):
            dur = m.group(1) + m.group(2)
        if dur not in out["must_contain_all"]:
            out["must_contain_all"].append(dur)

    # 7) Dates / années (2024, 2027)
    for m in RE_ANNEE.finditer(texte_clean):
        an = m.group(1)
        if an not in out["must_contain_all"]:
            out["must_contain_all"].append(an)

    # 8) Dates FR explicites
    for m in RE_DATE_FR.finditer(texte_clean):
        d = m.group(1)
        if d not in out["must_contain_all"]:
            out["must_contain_all"].append(d)

    # 9) Patterns négatifs — seuls les quotes APRÈS le marqueur "NE PAS ..." comptent
    neg_quoted: list[str] = []
    for phrase in re.split(r"[;.]\s*", texte):
        neg_pos = -1
        for rx in NEG_MARKERS:
            m = rx.search(phrase)
            if m:
                neg_pos = m.start()
                break
        if neg_pos < 0:
            continue
        # Extraire uniquement les quotes positionnées APRÈS le marqueur négatif
        for m in RE_QUOTED.finditer(phrase):
            if m.start() < neg_pos:
                continue
            val = (m.group(1) or m.group(2) or m.group(3) or "").strip().rstrip(",.;:!?")
            if val and val not in neg_quoted:
                neg_quoted.append(val)
    for v in neg_quoted:
        if v not in out["must_not_contain"]:
            out["must_not_contain"].append(v)
        out["must_contain_all"] = [x for x in out["must_contain_all"] if _normalize(x) != _normalize(v)]

    # Nettoyage : garde les nombres simples ≥ 1 char, exclut les vides.
    # Pour les chaînes alphabétiques : exige au moins 2 chars (évite faux pos. type 'l').
    def _keep(v: str) -> bool:
        v = v.strip()
        if not v:
            return False
        # Nombres : OK dès 1 char
        if re.match(r"^\d+[\d\s.,]*$", v):
            return True
        # Reste : ≥ 2 chars
        return len(v) >= 2
    out["must_contain_all"] = [v for v in out["must_contain_all"] if _keep(v)]
    # Déduplique en préservant l'ordre
    seen = set()
    deduped = []
    for v in out["must_contain_all"]:
        key = _normalize(v)
        if key not in seen:
            seen.add(key)
            deduped.append(v)
    out["must_contain_all"] = deduped

    out["parseable"] = bool(
        out["must_contain_all"] or out["must_contain_any"] or out["must_not_contain"]
    )
    return out


# ── Appel de l'API /api/ask ───────────────────────────────────────────────
def ask_bot(base_url: str, question: str, module: str = "juridique") -> dict:
    """Appelle /api/ask et retourne {answer, status, latency_ms, error?}."""
    t0 = time.time()
    url = base_url.rstrip("/") + "/api/ask"
    payload = {"question": question, "module": module, "history": []}
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(
        url,
        data=body,
        headers={"Content-Type": "application/json"},
        method="POST",
    )
    try:
        with urllib.request.urlopen(req, timeout=TIMEOUT_S) as resp:
            raw = resp.read().decode("utf-8", errors="replace")
            status = resp.status
            data = json.loads(raw)
            return {
                "answer": data.get("answer", ""),
                "status": status,
                "latency_ms": int((time.time() - t0) * 1000),
                "module": data.get("module"),
                "niveau": data.get("niveau"),
                "theme": data.get("theme"),
            }
    except urllib.error.HTTPError as e:
        body_txt = e.read().decode("utf-8", errors="replace")
        return {
            "answer": "",
            "status": e.code,
            "latency_ms": int((time.time() - t0) * 1000),
            "error": f"HTTP {e.code} : {body_txt[:200]}",
        }
    except urllib.error.URLError as e:
        return {
            "answer": "",
            "status": 0,
            "latency_ms": int((time.time() - t0) * 1000),
            "error": f"Erreur réseau : {e.reason}",
        }
    except Exception as e:
        return {
            "answer": "",
            "status": 0,
            "latency_ms": int((time.time() - t0) * 1000),
            "error": f"{type(e).__name__}: {e}",
        }


# ── Évaluation d'un cas ───────────────────────────────────────────────────
def evaluate_criteria(answer: str, criteria: dict) -> tuple[str, list[str]]:
    """Retourne ('OK' | 'KO' | 'MANUEL', details).

    - OK : tous les contrôles automatiques passent.
    - KO : au moins un contrôle échoue.
    - MANUEL : le critère n'a pas été parsé automatiquement.
    """
    details = []
    if not criteria["parseable"]:
        return "MANUEL", ["Critère non parsable automatiquement → vérification manuelle."]

    ok = True
    for needle in criteria["must_contain_all"]:
        if _contains(answer, needle):
            details.append(f"✓ contient '{needle}'")
        else:
            details.append(f"✗ MANQUANT : '{needle}'")
            ok = False

    for group in criteria["must_contain_any"]:
        if any(_contains(answer, g) for g in group):
            matched = next(g for g in group if _contains(answer, g))
            details.append(f"✓ contient au moins une option ({matched})")
        else:
            details.append(f"✗ MANQUE tout le groupe : {group}")
            ok = False

    for needle in criteria["must_not_contain"]:
        if _contains(answer, needle):
            details.append(f"✗ PRÉSENCE INTERDITE : '{needle}'")
            ok = False
        else:
            details.append(f"✓ absent : '{needle}'")

    return ("OK" if ok else "KO"), details


# ── Lecture du plan xlsx ──────────────────────────────────────────────────
def load_cases(xlsx_path: Path, sheet_name: str) -> list[dict]:
    """Charge la feuille 'Exactitude données' et retourne une liste de dicts."""
    if not xlsx_path.exists():
        raise FileNotFoundError(f"Fichier introuvable : {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(
            f"Feuille '{sheet_name}' absente. Feuilles : {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    # Trouver la ligne d'en-tête (cherche 'ID' en colonne A)
    header_row = None
    for r in range(1, min(ws.max_row, 10) + 1):
        val = ws.cell(row=r, column=1).value
        if val and str(val).strip().upper() == "ID":
            header_row = r
            break
    if header_row is None:
        raise ValueError("Ligne d'en-tête 'ID' introuvable.")

    headers = [ws.cell(row=header_row, column=c).value for c in range(1, ws.max_column + 1)]
    # Map canonique
    col_idx: dict[str, int] = {}
    for i, h in enumerate(headers):
        if not h:
            continue
        key = _normalize(str(h))
        col_idx[key] = i + 1

    def _col(name: str) -> int | None:
        return col_idx.get(_normalize(name))

    cases = []
    for r in range(header_row + 1, ws.max_row + 1):
        row_id = ws.cell(row=r, column=1).value
        if not row_id:
            continue
        cases.append({
            "row": r,
            "id": str(row_id).strip(),
            "categorie": str(ws.cell(row=r, column=_col("Catégorie") or 2).value or "").strip(),
            "donnee": str(ws.cell(row=r, column=_col("Donnée à vérifier") or 3).value or "").strip(),
            "reference": str(ws.cell(row=r, column=_col("Valeur de référence (source)") or 4).value or "").strip(),
            "question": str(ws.cell(row=r, column=_col("Question à poser au bot") or 5).value or "").strip(),
            "critere": str(ws.cell(row=r, column=_col("Critère de validation (réponse attendue)") or 6).value or "").strip(),
            "frequence": str(ws.cell(row=r, column=_col("Fréquence de contrôle") or 7).value or "").strip(),
            "col_derniere_verif": _col("Dernière vérif."),
            "col_statut": _col("Statut"),
        })
    return cases, wb, ws


# ── Rapport / CLI ──────────────────────────────────────────────────────────
def format_report(results: list[dict]) -> str:
    ok = sum(1 for r in results if r["verdict"] == "OK")
    ko = sum(1 for r in results if r["verdict"] == "KO")
    manuel = sum(1 for r in results if r["verdict"] == "MANUEL")
    err = sum(1 for r in results if r["verdict"] == "ERREUR")
    total = len(results)

    lines = []
    lines.append("=" * 72)
    lines.append(f"RAPPORT EXACTITUDE DONNÉES — {datetime.now(timezone.utc).isoformat(timespec='seconds')}")
    lines.append("=" * 72)
    lines.append(f"Total : {total}  |  ✓ OK : {ok}  |  ✗ KO : {ko}  |  ⚠ MANUEL : {manuel}  |  ☠ ERREUR : {err}")
    lines.append("")

    for r in results:
        icon = {"OK": "✓", "KO": "✗", "MANUEL": "⚠", "ERREUR": "☠"}[r["verdict"]]
        lines.append(f"{icon} [{r['id']}] {r['categorie']} — {r['donnee']}")
        if r["verdict"] == "KO":
            for d in r.get("details", []):
                if d.startswith("✗"):
                    lines.append(f"    {d}")
        elif r["verdict"] == "ERREUR":
            lines.append(f"    {r.get('error', 'erreur inconnue')}")
        elif r["verdict"] == "MANUEL":
            lines.append(f"    Réponse bot : {r.get('answer', '')[:120]}...")
    lines.append("")
    lines.append("=" * 72)
    lines.append(f"Taux de succès automatique : {ok}/{total - manuel - err} cas vérifiables ({100 * ok / max(1, total - manuel - err):.1f}%)")
    return "\n".join(lines)


def main() -> int:
    parser = argparse.ArgumentParser(description="Tests d'exactitude données — Chatbot ELISFA")
    parser.add_argument("--url", default=DEFAULT_URL, help=f"URL racine du bot (défaut : {DEFAULT_URL})")
    parser.add_argument("--xlsx", default=str(DEFAULT_XLSX), help="Chemin du plan xlsx")
    parser.add_argument("--sheet", default=DEFAULT_SHEET, help="Nom de la feuille")
    parser.add_argument("--filter", default=None, help="Filtre par ID ou catégorie (substring, insensible à la casse)")
    parser.add_argument("--limit", type=int, default=0, help="Nombre max de cas à tester (0 = tous)")
    parser.add_argument("--report", default=None, help="Chemin d'un rapport JSON à écrire")
    parser.add_argument("--update-xlsx", action="store_true", help="Met à jour les colonnes Statut + Dernière vérif. dans le xlsx")
    parser.add_argument("--pause", type=float, default=PAUSE_BETWEEN_CALLS_S, help="Pause entre appels (s)")
    parser.add_argument("--skip-manuel", action="store_true", help="Saute les cas dont le critère n'est pas auto-parseable (économie API)")
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    try:
        cases, wb, ws = load_cases(xlsx_path, args.sheet)
    except Exception as e:
        print(f"Erreur lecture xlsx : {e}", file=sys.stderr)
        return 2

    if args.filter:
        flt = args.filter.lower()
        cases = [c for c in cases if flt in c["id"].lower() or flt in c["categorie"].lower() or flt in c["donnee"].lower()]
    if args.skip_manuel:
        before = len(cases)
        cases = [c for c in cases if parse_criteria(c["critere"])["parseable"]]
        print(f"[i] --skip-manuel : {before - len(cases)} cas non-parseables ignorés")
    if args.limit > 0:
        cases = cases[: args.limit]

    print(f"[i] Tests sur {len(cases)} cas via {args.url}")
    print(f"[i] Xlsx : {xlsx_path}")
    if args.update_xlsx:
        print("[i] Mode --update-xlsx activé : les statuts seront réécrits dans le fichier.")
    print()

    results: list[dict] = []
    for i, case in enumerate(cases, 1):
        module = MODULE_PAR_CATEGORIE.get(case["categorie"], "juridique")
        if not case["question"]:
            results.append({
                **case,
                "verdict": "MANUEL",
                "details": ["Question absente → test manuel."],
                "answer": "",
                "latency_ms": 0,
            })
            print(f"[{i}/{len(cases)}] {case['id']}  ⚠ MANUEL (pas de question)")
            continue

        res = ask_bot(args.url, case["question"], module)
        time.sleep(args.pause)

        if res.get("error"):
            results.append({
                **case,
                "verdict": "ERREUR",
                "details": [res["error"]],
                "answer": "",
                "latency_ms": res["latency_ms"],
                "error": res["error"],
            })
            print(f"[{i}/{len(cases)}] {case['id']}  ☠ ERREUR  {res['error']}")
            continue

        criteria = parse_criteria(case["critere"])
        verdict, details = evaluate_criteria(res["answer"], criteria)
        results.append({
            **case,
            "verdict": verdict,
            "details": details,
            "answer": res["answer"][:500],
            "latency_ms": res["latency_ms"],
        })
        icon = {"OK": "✓", "KO": "✗", "MANUEL": "⚠"}[verdict]
        print(f"[{i}/{len(cases)}] {case['id']}  {icon} {verdict}  ({res['latency_ms']} ms)  {case['donnee'][:50]}")

    # Rapport texte
    print()
    print(format_report(results))

    # Rapport JSON
    if args.report:
        report = {
            "generated_at": datetime.now(timezone.utc).isoformat(timespec="seconds"),
            "url": args.url,
            "xlsx": str(xlsx_path),
            "total": len(results),
            "ok": sum(1 for r in results if r["verdict"] == "OK"),
            "ko": sum(1 for r in results if r["verdict"] == "KO"),
            "manuel": sum(1 for r in results if r["verdict"] == "MANUEL"),
            "erreur": sum(1 for r in results if r["verdict"] == "ERREUR"),
            "cases": [
                {k: v for k, v in r.items() if k not in ("col_derniere_verif", "col_statut", "row")}
                for r in results
            ],
        }
        Path(args.report).write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"[i] Rapport JSON : {args.report}")

    # Réécriture xlsx
    if args.update_xlsx:
        today = datetime.now().strftime("%Y-%m-%d")
        for r in results:
            if r.get("col_derniere_verif"):
                ws.cell(row=r["row"], column=r["col_derniere_verif"]).value = today
            if r.get("col_statut"):
                ws.cell(row=r["row"], column=r["col_statut"]).value = r["verdict"]
        wb.save(xlsx_path)
        print(f"[i] Xlsx mis à jour : {xlsx_path}")

    # Code retour : 1 si au moins un KO, 0 sinon
    return 1 if any(r["verdict"] == "KO" for r in results) else 0


if __name__ == "__main__":
    sys.exit(main())
